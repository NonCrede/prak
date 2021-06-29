import requests
import sys
from PyQt5 import QtCore, QtGui, QtWidgets, Qt, uic
from PyQt5.QtWidgets import *
import traceback
import pandas as pd
import time
import numpy as np
import openpyxl
import requests
import docx
from openpyxl import load_workbook
from datetime import datetime
import matplotlib.pyplot as plt
import os

def log_uncaught_exceptions(ex_cls, ex, tb): #error catcher
    text = '{}: {}:\n'.format(ex_cls.__name__, ex)
    text += ''.join(traceback.format_tb(tb))
    QtWidgets.QMessageBox.critical(None, 'Error', text)
    sys.exit()
sys.excepthook = log_uncaught_exceptions

mainwindow, _ = uic.loadUiType('kekw.ui')

workbook = load_workbook('test.xlsx')
column = workbook.active

Date1 = 0
Date2 = 0
Everything = {}
Vulner = []
Description = []
Soft = []
SoftType = []
VulnerClass = []
DateFound = []
VulnerLVL = []
Status = []
Exploit = []
ProtInfo = []
CWE = []
Measures = []
ID = []
Length = 0
Class = 0
Text = str()


class window(QMainWindow, mainwindow):
    def __init__(self):
        super(window, self).__init__()
        self.setupUi(self)
        self.initUI()


    def initUI(self):
        self.ShowDiagram.clicked.connect(lambda: self.diagram())
        self.UpdateBase.clicked.connect(lambda: self.updatefile())
        self.StartProgramm.clicked.connect(lambda: self.Result())
        ExploitList = ['Без фильтра', 'Данные уточняются', 'Существует', 'Существует в открытом доступе']
        self.ExploitNal.addItems(ExploitList)
        ExploitClassList = ['Без фильтра', 'Уязвимость архитектуры', 'Уязвимость кода', 'Уязвимость многофакторная']
        self.UyazClass.addItems(ExploitClassList)
        ExploitStatusList = ['Без фильтра', 'Подтверждена', 'Потенциальная']
        self.ExploitStat.addItems(ExploitStatusList)
        ExploitStatusInfo = ['Без фильтра', 'Информация об устранении отсутствует', 'Уязвимость устранена']
        self.InfoObUstr.addItems(ExploitStatusInfo)
        DangerLVLlist = ['Без фильтра', 'Критический', 'Высокий', 'Средний', 'Низкий']
        self.DangerLVL.addItems(DangerLVLlist)





    def updatefile(self):
        dls = "https://bdu.fstec.ru/files/documents/vullist.xlsx"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.106 Safari/537.36',
            'Referrer': 'https://bdu.fstec.ru/vul%27%7D'}
        resp = requests.get(dls, headers=headers)

        output = open('test.xlsx', 'wb')
        output.write(resp.content)
        output.close()


    def initial_D(self):
        Text = self.ProgrammName.text()
        Counter = 0
        global column
        for row in range(1, column.max_row):
            if Text in str(column.cell(row=row, column=5).value):
                Vulner.append(column.cell(row=row, column=2).value)
                Description.append(column.cell(row=row, column=3).value)
                Soft.append(column.cell(row=row, column=5).value)
                SoftType.append(column.cell(row=row, column=7).value)
                VulnerClass.append(column.cell(row=row, column=9).value)
                DateFound.append(column.cell(row=row, column=10).value)
                VulnerLVL.append(column.cell(row=row, column=13).value)
                Status.append(column.cell(row=row, column=15).value)
                Exploit.append(column.cell(row=row, column=16).value)
                ProtInfo.append(column.cell(row=row, column=17).value)
                CWE.append(column.cell(row=row, column=22).value)
                Measures.append(column.cell(row=row, column=14).value)
                ID.append(column.cell(row=row, column=2).value)
                Counter += 1
        for i in range(Counter):
            Everything[i] = (
                Vulner[i], Description[i], Soft[i], SoftType[i], VulnerClass[i], DateFound[i],
                VulnerLVL[i],
                Status[i], Exploit[i], ProtInfo[i], CWE[i], Measures[i], ID[i])

    def CountCheker(self, Count):
        Counter = 0
        for t in Count:
            Counter += 1
        return Counter


    def TranslateToDate(self, Date):
        Translate = "%d.%m.%Y"
        Dated = datetime.strptime(Date, Translate)
        return Dated

    '''
    def TranslatorTDate(Date):
        TranslatorTDate = "%d/%m/%Y"
        DateTranslator = datetime.strptime(Date, TranslatorTDate)
        return DateTranslator
        '''

    def FilterByClass(self, Length):
        ExploitClassList = self.UyazClass.currentText()
        if (ExploitClassList != 'Без фильтра'):
            for i in range(Length):
                try:
                    if ExploitClassList in str(Everything[i][4]):
                        continue
                    else:
                        del Everything[i]
                except KeyError:
                    continue


    def FilterByStatus(self, Length):
        ExploitStatusList = self.ExploitStat.currentText()
        if (ExploitStatusList != 'Без фильтра'):
            for i in range(Length):
                try:
                    if ExploitStatusList in str(Everything[i][7]):
                        continue
                    else:
                        del Everything[i]
                except KeyError:
                    continue

    def FilterByUSTRINFO(self):
        ExploitStatusInfo = self.ExploitStat.currentText()
        if (ExploitStatusInfo != 'Без фильтра'):
            for i in range(Length):
                try:
                    if ExploitStatusInfo in str(Everything[i][9]):
                        continue
                    else:
                        del Everything[i]
                except KeyError:
                    continue

    def FilterByDate(self, Date1, Date2, Length):
        global Everything
        Date1 = self.Date1Input.dateTime().toString('dd.MM.yyyy')
        Date2 = self.Date2Input.dateTime().toString('dd.MM.yyyy')
        for i in range(Length):
            try:
                if Everything[i][5] is None:
                    del Everything[i]
                    continue
                if (self.TranslateToDate(Everything[i][5]) < self.TranslateToDate(Date1)) | (
                        self.TranslateToDate(Everything[i][5]) > self.TranslateToDate(Date2)):
                    del Everything[i]
                    continue

            except KeyError:
                continue

    def FilterByCWE(self):
        CWEcode = str()
        CWEcode = self.spinBox.text()
        if (CWEcode != '0'):
            for i in range(Length):
                try:
                    if CWEcode in str(Everything[i][10]):
                        continue
                    else:
                        del Everything[i]
                except KeyError:
                    continue

    def FilterByExploit(self, Length):
        ExploitList = self.ExploitNal.currentText()
        if (ExploitList != 'Без фильтра'):
            for i in range(Length):
                try:
                    if ExploitList in str(Everything[i][8]):
                        continue
                    else:
                        del Everything[i]
                except KeyError:
                    continue


    def FilterByDangerLVL(self):
        DangerLVLlist = self.DangerLVL.currentText()
        if (DangerLVLlist != 'Без фильтра'):
            for i in range(Length):
                try:
                    if DangerLVLlist in str(Everything[i][6]):
                        continue
                    else:
                        del Everything[i]
                except KeyError:
                    continue

    def Report(self):
        Doc = self.docx.Document()
        Table = Doc.add_table(
            rows=1,
            cols=3)
        Table.style = 'Table Grid'
        Row = Table.rows[0].cells
        for i in range(Length):
            try:
                Row[0].text = str(Everything[i][0])
                Row[1].text = str(Everything[i][5])
                Row[2].text = str(Everything[i][6])
                Row = Table.add_row().cells
                continue
            except KeyError:
                continue
        Doc.save('Report.docx')

    def UyazCounter(self):
        CritCounter = 0
        HighCounter = 0
        MediumCounter = 0
        LowCounter = 0
        for i in Everything:
            try:
                if 'Критический' in str(Everything[i][6]):
                    CritCounter += 1
                    continue
                elif 'Высокий' in str(Everything[i][6]):
                    HighCounter += 1
                    continue
                elif 'Средний' in str(Everything[i][6]):
                    MediumCounter += 1
                    continue
                elif 'Низкий' in str(Everything[i][6]):
                    LowCounter += 1
                    continue
            except KeyError:
                continue
        return [CritCounter, HighCounter, MediumCounter, LowCounter]

    def diagram(self):
        index = ['Критический', 'Высокий', 'Средний', 'Низкий']
        values = self.UyazCounter()
        print(values)
        plt.bar(index, values)
        plt.title('Количественная диаграмма по уровню опасности')
        plt.savefig('graph.png', bbox_inches='tight')
        plt.show()

    def Result(self):
        self.initial_D()
        global Length
        Length = self.CountCheker(Everything)
        self.FilterByDate(Date1, Date2, Length)
        self.FilterByExploit(Length)
        self.FilterByClass(Length)
        self.FilterByStatus(Length)
        self.FilterByUSTRINFO()
        self.FilterByDangerLVL()
        self.FilterByCWE()



        for i in Everything:

            print(Everything[i][10])



if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    ui1 = window()
    ui1.show()
    sys.exit(app.exec_())
